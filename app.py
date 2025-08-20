from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import os
import json
import logging
from werkzeug.utils import secure_filename
from io import StringIO
import sys
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import fitz  # PyMuPDF
import numpy as np
import cv2
import easyocr
import re
from doctr.models import ocr_predictor
from collections import Counter
import time
from tqdm import tqdm
from typing import List, Dict, Tuple
import itertools
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuration
UPLOAD_FOLDER = 'Uploads'
ALLOWED_EXTENSIONS = {'pdf'}
RESULTS_FILE = 'optimization_results.json'

# Constants
MIN_PANEL_SIZE = 100
MAX_PANEL_SIZE = 600
STANDARD_PANEL_SIZES = [100, 200, 300, 400, 500, 600]

# Cache for storing previously computed panel combinations
panel_combinations_cache = {}

reader = easyocr.Reader(['en'])

# Setup
app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Logger setup
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# --- HELPER FUNCTIONS for TEXT CLASSIFICATION ---

def is_dimension(text: str) -> bool:
    """Checks if a string is a dimension (e.g., '2250X250')."""
    parts = text.split("X")
    if len(parts) != 2:
        return False
    return parts[0].isdigit() and parts[1].isdigit()

def is_part_label(text: str) -> bool:
    """
    Checks if a string is a valid part label.
    Rule: Must be alphanumeric, have both letters and digits, and be a reasonable length.
    """
    if len(text) > 10:
        return False
    if not text.isalnum():
        return False
    has_letter = any(c.isalpha() for c in text)
    has_digit = any(c.isdigit() for c in text)
    return has_letter and has_digit

def extract_text_from_image(image_path):
    """Extract text using the method from extraction_1.py"""
    image = cv2.imread(image_path)
    model = ocr_predictor(pretrained=True)
    results = model([cv2.cvtColor(image, cv2.COLOR_BGR2RGB)])
    ocr_data = results.export()["pages"][0]["blocks"]

    texts = []
    for block in ocr_data:
        for line in block["lines"]:
            for word in line["words"]:
                box = word["geometry"]
                text = word["value"].strip().upper()
                x_center = (box[0][0] + box[1][0]) / 2
                y_center = (box[0][1] + box[1][1]) / 2
                texts.append({"text": text, "x": x_center, "y": y_center})

    casting_label = None
    for t in texts:
        if re.search(r'CASTING\s*[—\-_]?\s*\d+', t['text'], re.IGNORECASE):
            casting_label = t['text']
            break

    dimensions = []
    labels = []
    for t in texts:
        if is_dimension(t['text']):
            dimensions.append(t)
        elif is_part_label(t['text']):
            labels.append(t)

    label_best_matches = {}
    for i, label in enumerate(labels):
        min_dist = float('inf')
        best_dim_idx = -1
        for j, dim in enumerate(dimensions):
            dist = np.hypot(label['x'] - dim['x'], label['y'] - dim['y'])
            if dist < min_dist:
                min_dist = dist
                best_dim_idx = j
        if best_dim_idx != -1:
            label_best_matches[i] = best_dim_idx

    dim_best_matches = {}
    for j, dim in enumerate(dimensions):
        min_dist = float('inf')
        best_label_idx = -1
        for i, label in enumerate(labels):
            dist = np.hypot(label['x'] - dim['x'], label['y'] - dim['y'])
            if dist < min_dist:
                min_dist = dist
                best_label_idx = i
        if best_label_idx != -1:
            dim_best_matches[j] = best_label_idx

    final_pairs = []
    for label_idx, dim_idx in label_best_matches.items():
        if dim_best_matches.get(dim_idx) == label_idx:
            label_text = labels[label_idx]['text']
            dim_text = dimensions[dim_idx]['text']
            final_pairs.append(f"{label_text} : {dim_text}")

    return final_pairs, casting_label

def extract_castings_from_pdf(pdf_path, dpi=400, min_area=5000):
    """
    Extract casting data from PDF and return structured data for panel optimization.
    Returns a list of dictionaries with casting information.
    """
    doc = fitz.open(pdf_path)
    page = doc[0]
    drawings = page.get_drawings()
    target_rectangles = []

    for drawing in drawings:
        for item in drawing["items"]:
            if item[0] == "re":
                rect = item[1]
                area = rect.width * rect.height
                if area >= min_area:
                    target_rectangles.append(rect)

    target_rectangles = sorted(target_rectangles, key=lambda r: r.width * r.height, reverse=True)

    castings_data = []
    
    for idx, rect in enumerate(target_rectangles):
        pix = page.get_pixmap(clip=rect, dpi=dpi)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
        if img.shape[2] == 4:
            img = img[:, :, :3]
        temp_path = f"casting_{idx+1}.png"
        cv2.imwrite(temp_path, cv2.cvtColor(img, cv2.COLOR_RGB2BGR))
        
        pairs, casting_label = extract_text_from_image(temp_path)
        
        if casting_label:
            casting_name = casting_label
        else:
            casting_name = f"CASTING_{idx+1}"
        
        shapes_data = {}
        for pair in pairs:
            if " : " in pair:
                label, dimension = pair.split(" : ")
                if is_dimension(dimension):
                    parts = dimension.split("X")
                    if len(parts) == 2 and parts[0].isdigit() and parts[1].isdigit():
                        length = int(parts[0])
                        width = int(parts[1])
                        shape_name = f"Shape_{label}"
                        shapes_data[shape_name] = {
                            "side1": length,
                            "side2": width,
                            "side3": length,
                            "side4": width
                        }
        
        if shapes_data:
            castings_data.append({
                "name": casting_name,
                "shapes": shapes_data
            })
        
        if os.path.exists(temp_path):
            os.remove(temp_path)
    
    doc.close()
    return castings_data

# --- PANEL OPTIMIZATION CLASSES AND FUNCTIONS ---

class Shape:
    def __init__(self, name: str, sides: List[int]):
        self.name = name
        self.sides = sides
        self.panel_layout = [[] for _ in range(len(sides))]
    
    def get_total_length(self) -> int:
        return sum(self.sides)
    
    def __str__(self) -> str:
        return f"Shape: {self.name}, Sides: {self.sides}"

class Casting:
    def __init__(self, name: str):
        self.name = name
        self.shapes = []
    
    def add_shape(self, shape: Shape) -> None:
        self.shapes.append(shape)
    
    def get_total_length(self) -> int:
        return sum(shape.get_total_length() for shape in self.shapes)
    
    def __str__(self) -> str:
        return f"Casting: {self.name}, Shapes: {len(self.shapes)}"

def analyze_castings(castings: List[Casting]) -> Dict:
    """
    Analyze all castings to identify common dimensions and optimal panel sizes.
    Returns information about preferred panel sizes for optimization.
    """
    length_counts = {}
    common_divisors = {}
    
    all_lengths = []
    for casting in castings:
        for shape in casting.shapes:
            all_lengths.extend(shape.sides)
    
    for length in all_lengths:
        length_counts[length] = length_counts.get(length, 0) + 1
    
    for panel_size in sorted(STANDARD_PANEL_SIZES, reverse=True):
        divisible_count = 0
        total_panels = 0
        
        for length in all_lengths:
            if length % panel_size == 0:
                divisible_count += 1
                total_panels += length // panel_size
            elif length % panel_size <= MIN_PANEL_SIZE and length >= panel_size:
                divisible_count += 0.5
                total_panels += length // panel_size
        
        efficiency = divisible_count / len(all_lengths) if all_lengths else 0
        common_divisors[panel_size] = {
            "efficiency": efficiency,
            "divisible_count": divisible_count,
            "total_panels": total_panels
        }
    
    panel_efficiency = sorted(
        [(size, data["efficiency"], data["total_panels"]) 
         for size, data in common_divisors.items()],
        key=lambda x: (-x[1], x[2])
    )
    
    preferred_sizes = [size for size, efficiency, _ in panel_efficiency if efficiency > 0.3]
    
    if len(preferred_sizes) < 2:
        for size in sorted(STANDARD_PANEL_SIZES, reverse=True):
            if size not in preferred_sizes:
                preferred_sizes.append(size)
            if len(preferred_sizes) >= 3:
                break
    
    return {
        "length_counts": length_counts,
        "preferred_sizes": preferred_sizes,
        "panel_efficiency": panel_efficiency
    }

def get_possible_panels(length: int) -> List[List[int]]:
    """
    Generate optimal panel combinations for a given length.
    Ensures all panels are within the valid size range (MIN_PANEL_SIZE to MAX_PANEL_SIZE).
    """
    if length in panel_combinations_cache:
        return panel_combinations_cache[length]
    
    valid_panels = []
    standard_sizes = sorted(STANDARD_PANEL_SIZES, reverse=True)
    
    if length < MIN_PANEL_SIZE:
        valid_panels.append([MIN_PANEL_SIZE])
        panel_combinations_cache[length] = valid_panels
        return valid_panels
    
    if length >= MAX_PANEL_SIZE:
        num_max_panels = length // MAX_PANEL_SIZE
        remaining = length - (num_max_panels * MAX_PANEL_SIZE)
        
        if remaining == 0:
            valid_panels.append([MAX_PANEL_SIZE] * num_max_panels)
        elif remaining >= MIN_PANEL_SIZE:
            valid_panels.append([MAX_PANEL_SIZE] * num_max_panels + [remaining])
        else:
            adjusted_length = MAX_PANEL_SIZE + remaining
            for r in range(1, 3):
                for combo in itertools.combinations_with_replacement(standard_sizes, r):
                    if sum(combo) == adjusted_length and all(p >= MIN_PANEL_SIZE for p in combo):
                        valid_panels.append([MAX_PANEL_SIZE] * (num_max_panels - 1) + list(combo))
            if not any(p for p in valid_panels if sum(p) == length) and adjusted_length >= MIN_PANEL_SIZE:
                valid_panels.append([MAX_PANEL_SIZE] * (num_max_panels - 1) + [adjusted_length])
    
    for r in range(1, min(8, length // MIN_PANEL_SIZE + 1)):
        for combo in itertools.combinations_with_replacement(standard_sizes, r):
            if sum(combo) == length:
                valid_panels.append(list(combo))
    
    for size1 in standard_sizes:
        if size1 <= length:
            max_count1 = min(length // size1, 8)
            for count1 in range(1, max_count1 + 1):
                remaining1 = length - (size1 * count1)
                if remaining1 == 0:
                    valid_panels.append([size1] * count1)
                elif remaining1 >= MIN_PANEL_SIZE:
                    for size2 in standard_sizes:
                        if size2 <= remaining1:
                            if remaining1 % size2 == 0:
                                count2 = remaining1 // size2
                                valid_panels.append([size1] * count1 + [size2] * count2)
                            elif remaining1 > size2 and (remaining1 % size2) >= MIN_PANEL_SIZE:
                                count2 = remaining1 // size2
                                last_panel = remaining1 - (size2 * count2)
                                if last_panel >= MIN_PANEL_SIZE:
                                    valid_panels.append([size1] * count1 + [size2] * count2 + [last_panel])
    
    if length <= MAX_PANEL_SIZE and length >= MIN_PANEL_SIZE:
        if length in STANDARD_PANEL_SIZES:
            valid_panels.append([length])
        else:
            valid_panels.append([length])
    
    valid_panels = [
        combo for combo in valid_panels 
        if all(MIN_PANEL_SIZE <= p <= MAX_PANEL_SIZE for p in combo) and sum(combo) == length
    ]
    
    verified_panels = []
    seen = set()
    for combo in valid_panels:
        combo_tuple = tuple(sorted(combo))
        if combo_tuple not in seen:
            verified_panels.append(combo)
            seen.add(combo_tuple)
    
    sorted_panels = sorted(
        verified_panels,
        key=lambda x: (
            sum(0 if p in STANDARD_PANEL_SIZES else 1 for p in x),
            len(x),
            -sum(p for p in x) / len(x) if x else 0,
            -sum(1 for p in x if p == MAX_PANEL_SIZE)
        )
    )
    
    if not sorted_panels and length > 0:
        if length < MIN_PANEL_SIZE:
            sorted_panels = [[MIN_PANEL_SIZE]]
        elif length <= MAX_PANEL_SIZE:
            sorted_panels = [[length]]
        else:
            max_count = length // MAX_PANEL_SIZE
            remaining = length % MAX_PANEL_SIZE
            if remaining >= MIN_PANEL_SIZE:
                sorted_panels = [[MAX_PANEL_SIZE] * max_count + [remaining]]
            else:
                adjusted = [MAX_PANEL_SIZE] * (max_count - 1)
                remaining_length = MAX_PANEL_SIZE + remaining
                for size in sorted(standard_sizes, reverse=True):
                    if remaining_length >= size + MIN_PANEL_SIZE:
                        final_remainder = remaining_length - size
                        if final_remainder >= MIN_PANEL_SIZE:
                            sorted_panels = [adjusted + [size, final_remainder]]
                            break
                if not sorted_panels:
                    if (max_count-1) * MAX_PANEL_SIZE >= length - MIN_PANEL_SIZE:
                        remaining = length - (max_count-1) * MAX_PANEL_SIZE
                        sorted_panels = [[MAX_PANEL_SIZE] * (max_count-1) + [remaining]]
                    else:
                        min_count = (length + MIN_PANEL_SIZE - 1) // MIN_PANEL_SIZE
                        sorted_panels = [[MIN_PANEL_SIZE] * min_count]
    
    panel_combinations_cache[length] = sorted_panels
    return sorted_panels

def optimize_panels(castings: List[Casting], primary_idx: int) -> Dict:
    """
    Optimize panel layout with cumulative reuse: panels from all previous castings are available for reuse in each subsequent casting.
    """
    print("\nOptimizing panel layouts with cumulative reuse...")
    start_time = time.time()

    ordered_castings = castings
    panel_inventory = {}
    new_panels_added = {}
    panels_used_per_casting = []

    for cast_idx, casting in enumerate(ordered_castings):
        casting_panels_used = {}
        for shape in casting.shapes:
            for side_idx, side_length in enumerate(shape.sides):
                layouts = get_possible_panels(side_length)
                if not layouts:
                    continue
                layout = layouts[0]
                shape.panel_layout[side_idx] = layout.copy()
                for panel in layout:
                    if panel_inventory.get(panel, 0) > 0:
                        panel_inventory[panel] -= 1
                    else:
                        new_panels_added[panel] = new_panels_added.get(panel, 0) + 1
                    casting_panels_used[panel] = casting_panels_used.get(panel, 0) + 1
        for panel, count in casting_panels_used.items():
            panel_inventory[panel] = panel_inventory.get(panel, 0) + count
        panels_used_per_casting.append(casting_panels_used)

    elapsed_time = time.time() - start_time
    print(f"\nOptimization completed in {elapsed_time:.2f} seconds.")
    return {
        "new_panels_added": new_panels_added,
        "panels_used_per_casting": panels_used_per_casting,
        "panel_inventory": panel_inventory
    }

def print_results(castings: List[Casting], primary_idx: int, optimization_results: Dict) -> None:
    """Print the optimized panel layouts for all castings with detailed cumulative reuse analysis."""
    print(f"\nResults (Primary Casting: {castings[primary_idx].name})\n")

    panels_used_per_casting = optimization_results["panels_used_per_casting"]
    new_panels_added = optimization_results["new_panels_added"]
    panel_inventory = optimization_results["panel_inventory"]

    for i, casting in enumerate(castings):
        print(f"{'*' * 20} {casting.name} {'*' * 20}")
        print("PRIMARY" if i == primary_idx else f"SECONDARY #{i if i > primary_idx else i+1}")
        for shape in casting.shapes:
            print(f"\n  Shape: {shape.name}")
            for side_idx, side_length in enumerate(shape.sides):
                panels = shape.panel_layout[side_idx]
                print(f"    Side {side_idx+1} (Length: {side_length}): {panels}")
        print(f"  Panels used in this casting: {panels_used_per_casting[i]}")

    print("\n" + "=" * 50)
    print("PANEL USAGE SUMMARY")
    print("=" * 50)
    print(f"Total unique panel types used: {len(panel_inventory)}")
    print(f"Panel inventory after all castings: {panel_inventory}")
    print(f"Total new panels needed: {sum(new_panels_added.values())}")
    print(f"Breakdown of new panels needed: {new_panels_added}")
    
    print("\n" + "=" * 50)
    print("PANEL REQUIREMENTS BY CASTING")
    print("=" * 50)
    total_new_panels = 0
    for i, casting in enumerate(castings):
        if i == primary_idx:
            panels_needed = sum(panels_used_per_casting[i].values())
            print(f"{casting.name} (PRIMARY): {panels_needed} new panels needed")
            total_new_panels += panels_needed
        else:
            used = panels_used_per_casting[i]
            temp_inventory = {}
            for j in range(i):
                for panel, count in panels_used_per_casting[j].items():
                    temp_inventory[panel] = temp_inventory.get(panel, 0) + count
            new_needed = 0
            for panel, count in used.items():
                reused_count = min(temp_inventory.get(panel, 0), count)
                new_needed += count - reused_count
            print(f"{casting.name} (SECONDARY): {new_needed} new panels needed")
            total_new_panels += new_needed
    
    print(f"\nTotal new panels needed for entire project: {total_new_panels}")

    print("\n" + "=" * 50)
    print("CASTING PANEL REQUIREMENTS ANALYSIS")
    print("=" * 50)
    
    primary_used = panels_used_per_casting[primary_idx]
    print(f"\n{castings[primary_idx].name} (PRIMARY CASTING):")
    print(f"  Panels used: {primary_used}")
    print(f"  Reused from previous: {dict()}")
    print(f"  New panels needed: {primary_used}")
    
    cumulative_inventory = primary_used.copy()
    
    for i in range(len(castings)):
        if i == primary_idx:
            continue
        used = panels_used_per_casting[i]
        reused = {}
        new_needed = {}
        temp_inventory = cumulative_inventory.copy()
        
        for panel, count in used.items():
            reused_count = min(temp_inventory.get(panel, 0), count)
            reused[panel] = reused_count
            new_needed[panel] = count - reused_count
            temp_inventory[panel] = temp_inventory.get(panel, 0) - reused_count
        
        print(f"\n{castings[i].name} (SECONDARY #{i if i > primary_idx else i+1}):")
        print(f"  Panels used: {used}")
        print(f"  Reused from previous: {reused}")
        print(f"  New panels needed: {new_needed}")
        
        for panel, count in used.items():
            cumulative_inventory[panel] = cumulative_inventory.get(panel, 0) + count

    total_panels_used = 0
    total_panels_reused = 0
    
    for i in range(len(castings)):
        used = panels_used_per_casting[i]
        temp_inventory = {}
        for j in range(i):
            for panel, count in panels_used_per_casting[j].items():
                temp_inventory[panel] = temp_inventory.get(panel, 0) + count
        for panel, count in used.items():
            reused_count = min(temp_inventory.get(panel, 0), count)
            total_panels_used += count
            total_panels_reused += reused_count
    
    if total_panels_used > 0:
        efficiency = (total_panels_reused / total_panels_used) * 100
        print(f"\nOverall panel reuse efficiency: {efficiency:.1f}% ({total_panels_reused} of {total_panels_used} panels reused across all castings)")
        if len(castings) > 1:
            secondary_used = sum(sum(panels_used_per_casting[i].values()) for i in range(1, len(castings)))
            secondary_reused = 0
            for i in range(1, len(castings)):
                used = panels_used_per_casting[i]
                temp_inventory = {}
                for j in range(i):
                    for panel, count in panels_used_per_casting[j].items():
                        temp_inventory[panel] = temp_inventory.get(panel, 0) + count
                for panel, count in used.items():
                    reused_count = min(temp_inventory.get(panel, 0), count)
                    secondary_reused += reused_count
            if secondary_used > 0:
                secondary_efficiency = (secondary_reused / secondary_used) * 100
                print(f"Secondary casting panel reuse efficiency: {secondary_efficiency:.1f}% ({secondary_reused} of {secondary_used} panels reused)")
    else:
        print("\nPanel reuse efficiency: N/A (no castings)")

def convert_extracted_data_to_castings(castings_data: List[Dict]) -> List[Casting]:
    """Convert extracted casting data to Casting objects for optimization."""
    castings = []
    
    for casting_data in castings_data:
        casting = Casting(casting_data["name"])
        for shape_name, sides_data in casting_data["shapes"].items():
            sides = [length for _, length in sides_data.items()]
            shape = Shape(shape_name, sides)
            casting.add_shape(shape)
        castings.append(casting)
    
    return castings

def export_to_excel(panel_data):
    sorted_panels = sorted(panel_data, key=lambda x: x["width"], reverse=True)
    wb = Workbook()
    ws = wb.active

    ws.merge_cells('A1:H1')
    ws['A1'] = "COSMOS CONSTRUCTION MACHINERIES & EQUIPMENT PVT.LTD."
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:H2')
    ws['A2'] = "Plot No.E/307/5, Gat No.307, Nanekarwadi, Next to MINDA, Chakan, Pune. Phone -+91 86008 84281"
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:H3')
    ws['A3'] = "Email: aluformservice@cosmossales.com"
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A5:H5')
    ws['A5'] = "TOTAL CONSOLIDATED MATERIAL QUANTITY"
    ws['A5'].alignment = Alignment(horizontal='center')
    ws['A5'].font = Font(bold=True)

    ws.merge_cells('A6:H6')
    ws['A6'] = "SHUTTERING AREA STATEMENT"
    ws['A6'].alignment = Alignment(horizontal='center')
    ws['A6'].font = Font(bold=True)

    ws.merge_cells('A7:G7')
    ws['A7'] = "Project Name"
    ws['A7'].font = Font(bold=True)
    ws['A7'].alignment = Alignment(horizontal='center')

    ws['H7'] = datetime.now().strftime("%d.%m.%Y")
    ws['H7'].font = Font(bold=True)
    ws['H7'].alignment = Alignment(horizontal='right')

    table_headers = [
        "Sr. No.", "Panel Size (mm)", "Width (mm)", "Code", "Length (mm)",
        "Total No. Of Panel Quantity", "Unit Area of Panel in Sq.m", "Total Area in Sq.m of all Panels"
    ]
    ws.append(table_headers)
    for col in range(1, len(table_headers) + 1):
        ws.cell(row=8, column=col).font = Font(bold=True)
        ws.cell(row=8, column=col).alignment = Alignment(horizontal='center')

    total_area_sum = 0
    total_qty_sum = 0
    for idx, panel in enumerate(sorted_panels, start=1):
        width = panel["width"]
        length = panel["length"]
        code = panel["code"]
        qty = panel["quantity"]
        panel_size_str = f"{width}{code}{length}"
        unit_area = round((width * length) / 1_000_000, 3)
        total_area = round(unit_area * qty, 2)
        total_area_sum += total_area
        total_qty_sum += qty
        ws.append([
            idx,
            panel_size_str,
            width,
            code,
            length,
            qty,
            unit_area,
            total_area
        ])

    ws.append(["", "", "", "", "", total_qty_sum])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)
    ws.append(["", "", "", "", "", "", "Total Shuttering Area in Sq.m", round(total_area_sum, 2)])
    ws.cell(row=ws.max_row, column=7).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=8).font = Font(bold=True)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20

    wb.save("consolidated_list.xlsx")
    print("✅ Excel file 'consolidated_list.xlsx' generated successfully.")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/extract-castings', methods=['POST'])
def extract_castings():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            file.save(file_path)
            castings_data = extract_castings_from_pdf(file_path)
            os.remove(file_path)
            return jsonify(castings_data), 200
        except Exception as e:
            logging.exception("Failed to process PDF")
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'error': f"Failed to process PDF: {str(e)}"}), 500
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/api/optimize', methods=['POST']) 
def optimize():
    try:
        data = request.get_json()
        if not data or 'castings' not in data or 'primaryIdx' not in data:
            return jsonify({'error': 'Invalid request data'}), 400

        castings = convert_extracted_data_to_castings(data['castings'])
        primary_idx = int(data['primaryIdx'])

        if primary_idx < 0 or primary_idx >= len(castings):
            return jsonify({'error': f"Invalid primaryIdx: {primary_idx}"}), 400

        # Analyze castings
        analysis_results = analyze_castings(castings)

        # Perform optimization
        optimization_results = optimize_panels(castings, primary_idx)

        # Capture print_results output
        original_stdout = sys.stdout
        sys.stdout = buffer = StringIO()
        print_results(castings, primary_idx, optimization_results)
        sys.stdout = original_stdout
        text_summary = buffer.getvalue()

        # Prepare panel_data for Excel export
        new_panels_added = optimization_results["new_panels_added"]
        panel_data = [
            {
                "width": panel_size,
                "length": 2400,
                "code": "WS",
                "quantity": quantity
            }
            for panel_size, quantity in new_panels_added.items()
        ]

        # Export to Excel
        export_to_excel(panel_data)

        # Prepare results for JSON
        results = {
            'castings': [],
            'optimization_summary': {
                'total_castings': len(castings),
                'primary_casting': castings[primary_idx].name,
                'panel_sizes_used': list(set(
                    panel for casting in castings
                    for shape in casting.shapes
                    for panels in shape.panel_layout
                    for panel in panels
                )),
                'total_new_panels_needed': sum(new_panels_added.values()),
                'panel_reuse_analysis': {
                    'primary_casting': {
                        'name': castings[primary_idx].name,
                        'panels_used': optimization_results["panels_used_per_casting"][primary_idx],
                        'new_panels_needed': optimization_results["panels_used_per_casting"][primary_idx]
                    },
                    'secondary_castings': []
                },
                'panel_analysis': {
                    'length_counts': analysis_results['length_counts'],
                    'preferred_sizes': analysis_results['preferred_sizes'],
                    'panel_efficiency': analysis_results['panel_efficiency']
                },
                'text_summary': text_summary
            }
        }

        # Add secondary casting analysis
        panels_used_per_casting = optimization_results["panels_used_per_casting"]
        cumulative_inventory = panels_used_per_casting[primary_idx].copy()
        for i in range(len(castings)):
            if i == primary_idx:
                continue
            used = panels_used_per_casting[i]
            reused = {}
            new_needed = {}
            temp_inventory = cumulative_inventory.copy()
            for panel, count in used.items():
                reused_count = min(temp_inventory.get(panel, 0), count)
                reused[panel] = reused_count
                new_needed[panel] = count - reused_count
                temp_inventory[panel] = temp_inventory.get(panel, 0) - reused_count
            results["optimization_summary"]["panel_reuse_analysis"]["secondary_castings"].append({
                "name": castings[i].name,
                "panels_used": used,
                "reused_from_previous": reused,
                "new_panels_needed": new_needed
            })
            for panel, count in used.items():
                cumulative_inventory[panel] = cumulative_inventory.get(panel, 0) + count

        for casting in castings:
            casting_data = {
                'name': casting.name,
                'shapes': []
            }
            for shape in casting.shapes:
                shape_data = {
                    'name': shape.name,
                    'sides': shape.sides,
                    'panel_layouts': shape.panel_layout
                }
                casting_data['shapes'].append(shape_data)
            results['castings'].append(casting_data)

        # Save results to file
        with open(RESULTS_FILE, 'w') as f:
            json.dump(results, f, indent=2)

        return jsonify(results), 200

    except ValueError as ve:
        logging.error(f"ValueError: {str(ve)}")
        return jsonify({'error': f"Invalid data: {str(ve)}"}), 400
    except Exception as e:
        logging.exception("Optimization failed")
        return jsonify({'error': f"Optimization failed: {str(e)}"}), 500

@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    excel_file = "consolidated_list.xlsx"
    try:
        if os.path.exists(excel_file):
            return send_file(excel_file, as_attachment=True, download_name='consolidated_list.xlsx')
        else:
            return jsonify({'error': 'Excel file not found. Please run optimization first.'}), 404
    except Exception as e:
        logging.exception("Failed to download Excel file")
        return jsonify({'error': f"Failed to download Excel file: {str(e)}"}), 500

if __name__ == '__main__':
    logging.info("Starting Flask server on http://localhost:5000")
    app.run(debug=True, port=5000)