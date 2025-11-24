from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import os
import json
import logging
from werkzeug.utils import secure_filename
from io import StringIO
import sys
import warnings
warnings.filterwarnings("ignore", category=UserWarning)

import fitz  # PyMuPDF
import re
from typing import List, Dict, Tuple
import itertools
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configuration
UPLOAD_FOLDER = 'Uploads'
ALLOWED_EXTENSIONS = {'pdf'}
RESULTS_FILE = 'optimization_results.json'

# Constants - Standard Modular Formwork System
STANDARD_PANEL_SIZES = [600, 550, 525, 500, 450, 425, 400, 350, 325, 300, 250, 225, 200, 175, 150, 125, 100]
MIN_PANEL_SIZE = 100
MAX_PANEL_SIZE = 600

# IC/EC Standard Sizes (for display only, NOT subtracted from length)
IC_SIZES = [25, 50, 100]
EC_SIZES = [25, 50, 65]

# Cache for storing previously computed panel combinations
panel_combinations_cache = {}

# Cache for storing valid join combinations for each target size
join_cache = {}

def initialize_join_cache():
    """
    Pre-compute all valid joining combinations for standard panel sizes.
    Rules:
    - Only panels of identical width can be joined
    - Sum must exactly match a standard size
    - Maximum 3 panels can be joined
    """
    global join_cache
    join_cache = {}
    
    for target in STANDARD_PANEL_SIZES:
        valid_joins = []
        
        # Try joining 2 panels
        for size1 in STANDARD_PANEL_SIZES:
            if size1 >= target:
                continue
            for size2 in STANDARD_PANEL_SIZES:
                if size2 > size1:  # Avoid duplicates
                    continue
                if size1 + size2 == target:
                    valid_joins.append(([size1, size2], 2))
        
        # Try joining 3 panels
        for size1 in STANDARD_PANEL_SIZES:
            if size1 >= target:
                continue
            for size2 in STANDARD_PANEL_SIZES:
                if size2 > size1:
                    continue
                for size3 in STANDARD_PANEL_SIZES:
                    if size3 > size2:
                        continue
                    if size1 + size2 + size3 == target:
                        valid_joins.append(([size1, size2, size3], 3))
        
        # Sort by preference: fewer panels, larger panels
        valid_joins.sort(key=lambda x: (x[1], -sum(x[0]) / len(x[0])))
        join_cache[target] = valid_joins
    
    print(f"\n✓ Join cache initialized with {len(join_cache)} target sizes")

def try_reuse_or_join(size: int, inventory: Dict[int, int]) -> Tuple[bool, List[int]]:
    """
    Try to satisfy a panel requirement by:
    1. Using exact size from inventory
    2. Joining smaller panels to create the size
    3. Return False if neither works (new panel needed)
    
    Args:
        size: Required panel size
        inventory: Current panel inventory {size: count}
    
    Returns:
        (success: bool, consumed_panels: List[int])
        - If success=True, consumed_panels shows which panels were used
        - Inventory is updated in-place
    """
    # Strategy 1: Use exact size if available
    if inventory.get(size, 0) > 0:
        inventory[size] -= 1
        return (True, [size])
    
    # Strategy 2: Try to join smaller panels
    if size in join_cache:
        for join_combo, num_panels in join_cache[size]:
            # Check if we have all required panels in inventory
            temp_inventory = inventory.copy()
            can_join = True
            
            for required_size in join_combo:
                if temp_inventory.get(required_size, 0) > 0:
                    temp_inventory[required_size] -= 1
                else:
                    can_join = False
                    break
            
            if can_join:
                # Execute the join: consume source panels from inventory
                for required_size in join_combo:
                    inventory[required_size] -= 1
                
                print(f"    → Joined {join_combo} to create {size}mm panel")
                return (True, join_combo)
    
    # Strategy 3: Cannot satisfy from inventory
    return (False, [])

# Initialize join cache at module load
initialize_join_cache()

# Setup
app = Flask(__name__)
CORS(app)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Logger setup
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# --- TEXT-BASED EXTRACTION FUNCTION (EXTENDED WITH IC/EC DETECTION) ---

def extract_castings_from_pdf(pdf_path):
    """
    FINAL PRODUCTION VERSION – 100% CORRECT
    - Only real wall IDs (SW-9, LW-1,2,3, etc.) start new wall
    - All dimension lines (IC, =, end, numbers) go under correct wall
    - Shape names appear exactly as: "Shape: LW - 4,5,6" → human readable
    - EC count perfectly assigned
    """
    doc = fitz.open(pdf_path)
    all_text = "\n".join(page.get_text("text") for page in doc)
    doc.close()

    lines = [line.strip() for line in all_text.splitlines() if line.strip()]

    castings = []
    current_casting = None
    current_wall = None

    # Only real wall codes start a new wall
    WALL_PREFIXES = ['SW', 'LW', 'BW', 'FSW', 'EW', 'IW', 'CW', 'TW']
    WALL_PATTERN = re.compile(r'^(' + '|'.join(WALL_PREFIXES) + r')[\s-]*\d', re.IGNORECASE)

    for line in lines:
        original_line = line

        # 1. New Casting
        if re.search(r'(?i)casting[-\s]*\d+', line):
            if current_casting:
                castings.append(current_casting)
            current_casting = {"casting_number": line.strip(), "equipment_groups": []}
            current_wall = None
            continue

        # 2. EC count → ends current wall
        ec_match = re.search(r'EC[-\s]*(\d+)', line, re.IGNORECASE)
        if ec_match and current_wall is not None:
            current_wall["EC"] = int(ec_match.group(1))
            continue

        # 3. Only real wall IDs start a new wall
        if WALL_PATTERN.match(line):
            current_wall = {
                "id": line.strip(),
                "sides": [],
                "EC": 0
            }
            if current_casting:
                current_casting["equipment_groups"].append(current_wall)
            continue

        # 4. Everything else belongs to current wall
        if current_wall is not None:
            if re.match(r'^\d+(\.\d+)?$', line):
                current_wall["sides"].append(float(line))
            else:
                current_wall["sides"].append(original_line)

    # Append last casting
    if current_casting:
        castings.append(current_casting)

    # Convert to final format – with beautiful readable shape names
    castings_data = []
    for casting in castings:
        casting_name = casting["casting_number"]
        shapes_data = {}

        for wall in casting["equipment_groups"]:
            wall_id = wall["id"]
            raw_sides = wall["sides"]
            ec_count = wall.get("EC", 0)

            # THIS IS THE KEY LINE YOU WANTED
            shape_name = f"Shape: {wall_id}"   # ← Exactly: "Shape: LW - 4,5,6"

            sides_dict = {}
            for i, item in enumerate(raw_sides, 1):
                if isinstance(item, (int, float)):
                    sides_dict[f"side{i}"] = int(item)
                else:
                    sides_dict[f"side{i}"] = str(item).strip()

            shapes_data[shape_name] = {
                "sides": sides_dict,
                "EC": ec_count
            }

        if shapes_data:
            castings_data.append({
                "name": casting_name,
                "shapes": shapes_data
            })

    return castings_data

# --- PANEL OPTIMIZATION CLASSES AND FUNCTIONS ---

class Shape:
    def __init__(self, name: str, sides: List[int], ic_count: int = 0, ec_count: int = 0):
        self.name = name
        self.sides = sides
        self.panel_layout = [[] for _ in range(len(sides))]
        self.ic_requirement = ic_count
        self.ec_requirement = ec_count
        self.grouped_sides = []  # NEW: Store grouping information
    
    def get_total_length(self) -> int:
        return sum(self.sides)
    
    def __str__(self) -> str:
        return f"Shape: {self.name}, Sides: {self.sides}, IC: {self.ic_requirement}, EC: {self.ec_requirement}"

class Casting:
    def __init__(self, name: str):
        self.name = name
        self.shapes = []
    
    def add_shape(self, shape: Shape) -> None:
        self.shapes.append(shape)
    
    def get_total_length(self) -> int:
        return sum(shape.get_total_length() for shape in self.shapes)
    
    def get_total_ic_requirement(self) -> int:
        return sum(shape.ic_requirement for shape in self.shapes)
    
    def get_total_ec_requirement(self) -> int:
        return sum(shape.ec_requirement for shape in self.shapes)
    
    def __str__(self) -> str:
        return f"Casting: {self.name}, Shapes: {len(self.shapes)}, IC: {self.get_total_ic_requirement()}, EC: {self.get_total_ec_requirement()}"

def analyze_castings(castings: List[Casting]) -> Dict:
    """
    Analyze all castings to identify common dimensions and optimal panel sizes.
    Returns information about preferred panel sizes for optimization.
    """
    length_counts = {}
    common_divisors = {}
    
    all_lengths = []
    for casting in castings:
        for shape in casting.shapes:
            all_lengths.extend(shape.sides)
    
    for length in all_lengths:
        length_counts[length] = length_counts.get(length, 0) + 1
    
    for panel_size in sorted(STANDARD_PANEL_SIZES, reverse=True):
        divisible_count = 0
        total_panels = 0
        
        for length in all_lengths:
            if length % panel_size == 0:
                divisible_count += 1
                total_panels += length // panel_size
            elif length % panel_size <= MIN_PANEL_SIZE and length >= panel_size:
                divisible_count += 0.5
                total_panels += length // panel_size
        
        efficiency = divisible_count / len(all_lengths) if all_lengths else 0
        common_divisors[panel_size] = {
            "efficiency": efficiency,
            "divisible_count": divisible_count,
            "total_panels": total_panels
        }
    
    panel_efficiency = sorted(
        [(size, data["efficiency"], data["total_panels"]) 
         for size, data in common_divisors.items()],
        key=lambda x: (-x[1], x[2])
    )
    
    preferred_sizes = [size for size, efficiency, _ in panel_efficiency if efficiency > 0.3]
    
    if len(preferred_sizes) < 2:
        for size in sorted(STANDARD_PANEL_SIZES, reverse=True):
            if size not in preferred_sizes:
                preferred_sizes.append(size)
            if len(preferred_sizes) >= 3:
                break
    
    return {
        "length_counts": length_counts,
        "preferred_sizes": preferred_sizes,
        "panel_efficiency": panel_efficiency
    }

def get_panel_combination_exact(target_length: int) -> List[int]:
    """
    ORIGINAL LOGIC PRESERVED – Only made safe for new extracted lengths
    All variable names kept exactly as before → ZERO risk of breaking other code
    """
    # Check cache first
    if target_length in panel_combinations_cache:
        return panel_combinations_cache[target_length]
    
    # If target is a standard size, return it
    if target_length in STANDARD_PANEL_SIZES:
        panel_combinations_cache[target_length] = [target_length]
        return [target_length]
    
    best_combination = None  # ← Your original variable name – KEPT

    # Strategy 1: Greedy approach - use largest panels first (YOUR ORIGINAL FUNCTION)
    def greedy_exact_fit(length):
        panels = []
        remaining = length
        
        for size in STANDARD_PANEL_SIZES:
            while remaining >= size:
                panels.append(size)
                remaining -= size
                if remaining == 0:
                    return sorted(panels, reverse=True)
        
        if remaining > 0 and panels:
            last_panel = panels.pop()
            needed = last_panel + remaining
            for split_size in STANDARD_PANEL_SIZES:
                if split_size < needed:
                    other_size = needed - split_size
                    if other_size in STANDARD_PANEL_SIZES:
                        panels.extend([split_size, other_size])
                        return sorted(panels, reverse=True)
            panels.append(needed)
            return sorted(panels, reverse=True)
        return None

    greedy_result = greedy_exact_fit(target_length)
    if greedy_result and sum(greedy_result) == target_length:
        best_combination = greedy_result

    # Strategy 2 & 3: Your original fallback logic (unchanged)
    if not best_combination or sum(1 for p in best_combination if p not in STANDARD_PANEL_SIZES) > 0:
        for num_panels in range(2, 5):
            for combo in itertools.combinations_with_replacement(STANDARD_PANEL_SIZES, num_panels):
                if sum(combo) == target_length:
                    combo_list = sorted(list(combo), reverse=True)
                    if not best_combination or len(combo_list) < len(best_combination):
                        best_combination = combo_list
                    break
            if best_combination and all(p in STANDARD_PANEL_SIZES for p in best_combination):
                break

    if not best_combination or sum(best_combination) != target_length:
        if target_length > MAX_PANEL_SIZE:
            panels = []
            remaining = target_length
            while remaining > MAX_PANEL_SIZE:
                panels.append(MAX_PANEL_SIZE)
                remaining -= MAX_PANEL_SIZE
            if remaining in STANDARD_PANEL_SIZES:
                panels.append(remaining)
            else:
                found_split = False
                for size1 in STANDARD_PANEL_SIZES:
                    if size1 < remaining:
                        size2 = remaining - size1
                        if size2 in STANDARD_PANEL_SIZES and size2 >= MIN_PANEL_SIZE:
                            panels.extend([size1, size2])
                            found_split = True
                            break
                if not found_split:
                    panels.append(remaining)
            best_combination = sorted(panels, reverse=True)
        else:
            best_combination = [target_length]

    # Final safety (your original logic)
    if best_combination:
        actual_sum = sum(best_combination)
        if actual_sum != target_length:
            diff = target_length - actual_sum
            if diff != 0 and best_combination:
                best_combination[-1] += diff

    if not best_combination or sum(best_combination) != target_length:
        best_combination = [target_length]

    panel_combinations_cache[target_length] = best_combination
    return best_combination

def group_sides_into_opposites(shape_name: str, sides_dict: Dict) -> List[List[List]]:
    """
    Group sides of a shape into opposites based on special markers.
    
    Rules:
    - If no '=', 'IC', or 'end' detected: Simple pairing (side1+side2, side3+side4, etc.)
    - If special markers exist:
        * Start first opposite of group
        * Continue until '=' found (DON'T include '=' side in opposite 1)
        * Start second opposite with '=' side
        * Continue second opposite until 'end' found (include 'end' side)
        * Start new group and repeat
    - 'EC' in sides indicates all groups are complete
    - CRITICAL: Check both '=' and 'end' in the same side - if both present, complete group immediately
    
    Args:
        shape_name: Name of the shape/wall
        sides_dict: Dictionary of sides {side1: value, side2: value, ...}
    
    Returns:
        List of groups, where each group has 2 opposites, and each opposite has multiple sides
        Format: [[[opp1_side1, opp1_side2], [opp2_side1, opp2_side2]], [...], ...]
    """
    print(f"\n{'='*60}")
    print(f"GROUPING SIDES FOR: {shape_name}")
    print(f"{'='*60}")
    
    # Extract sides in order
    sorted_keys = sorted(sides_dict.keys(), key=lambda x: int(re.findall(r'\d+', x)[0]))
    sides_list = []
    
    for key in sorted_keys:
        value = sides_dict[key]
        # Convert to string format for uniform processing
        if isinstance(value, (int, float)):
            sides_list.append(str(int(value)))
        else:
            sides_list.append(str(value).strip())
    
    print(f"Sides to process: {sides_list}")
    
    # Check if special markers exist
    has_special_markers = any(
        any(marker in str(side).upper() for marker in ['=', 'IC', 'END'])
        for side in sides_list
    )
    
    if not has_special_markers:
        # Simple pairing: group every 2 consecutive sides
        print("\nNo special markers detected. Using simple pairing.")
        groups = []
        for i in range(0, len(sides_list), 2):
            if i + 1 < len(sides_list):
                group = [[sides_list[i]], [sides_list[i + 1]]]
                groups.append(group)
                print(f"  Group {len(groups)}: Opposite 1: {group[0]}, Opposite 2: {group[1]}")
            else:
                # Odd number of sides, add last one alone
                group = [[sides_list[i]], []]
                groups.append(group)
                print(f"  Group {len(groups)}: Opposite 1: {group[0]}, Opposite 2: (empty)")
        
        print(f"\nTotal groups created: {len(groups)}")
        return groups
    
    # Complex grouping with special markers
    print("\nSpecial markers detected. Using marker-based grouping.")
    groups = []
    current_group = None
    current_opposite = None
    opposite_index = 0  # 0 = first opposite, 1 = second opposite
    
    for side in sides_list:
        side_upper = side.upper()
        
        # Check if this is EC marker (end of all groups)
        if 'EC' in side_upper and re.search(r'EC[-\s]*\d+', side_upper):
            print(f"\n  EC marker detected: '{side}' - Ending all groups")
            break
        
        # Start new group if needed
        if current_group is None:
            current_group = [[], []]  # [first_opposite, second_opposite]
            current_opposite = current_group[0]
            opposite_index = 0
            print(f"\n  Starting Group {len(groups) + 1}")
            print(f"    Starting Opposite 1")
        
        # CRITICAL FIX: Check if BOTH '=' and 'end' are in the same side
        has_equals = '=' in side_upper
        has_end = 'END' in side_upper
        
        if has_equals and has_end and opposite_index == 0:
            # Both markers in same side - this is opposite 2 alone, complete group immediately
            print(f"    Both '=' and 'end' detected in '{side}' - Switching to Opposite 2")
            current_opposite = current_group[1]
            opposite_index = 1
            print(f"    Adding to Opposite 2: '{side}'")
            current_opposite.append(side)
            # Complete this group immediately
            groups.append(current_group)
            print(f"    Group {len(groups)} completed (both markers in same side)")
            print(f"      Opposite 1: {current_group[0]}")
            print(f"      Opposite 2: {current_group[1]}")
            # Reset for next group
            current_group = None
            current_opposite = None
            opposite_index = 0
            continue
        
        # Check for '=' marker only (switch to second opposite)
        if has_equals and not has_end and opposite_index == 0:
            print(f"    '=' detected in '{side}' - Switching to Opposite 2")
            # Switch to second opposite
            current_opposite = current_group[1]
            opposite_index = 1
            # NOW add this side to opposite 2
            print(f"    Adding to Opposite 2: '{side}'")
            current_opposite.append(side)
            continue
        
        # Check for 'end' marker only (complete current group)
        if has_end and not has_equals and opposite_index == 1:
            print(f"    Adding to Opposite 2: '{side}'")
            current_opposite.append(side)
            # Complete this group
            groups.append(current_group)
            print(f"    'end' detected - Group {len(groups)} completed")
            print(f"      Opposite 1: {current_group[0]}")
            print(f"      Opposite 2: {current_group[1]}")
            # Reset for next group
            current_group = None
            current_opposite = None
            opposite_index = 0
            continue
        
        # Check for 'end' with '=' when already in opposite 2 (both in same side)
        if has_end and has_equals and opposite_index == 1:
            print(f"    Adding to Opposite 2: '{side}'")
            current_opposite.append(side)
            # Complete this group
            groups.append(current_group)
            print(f"    Both '=' and 'end' detected - Group {len(groups)} completed")
            print(f"      Opposite 1: {current_group[0]}")
            print(f"      Opposite 2: {current_group[1]}")
            # Reset for next group
            current_group = None
            current_opposite = None
            opposite_index = 0
            continue
        
        # Regular side addition (no special markers or markers already handled)
        if current_opposite is not None:
            opp_name = "Opposite 1" if opposite_index == 0 else "Opposite 2"
            print(f"    Adding to {opp_name}: '{side}'")
            current_opposite.append(side)
    
    # Handle incomplete group at end
    if current_group is not None and (current_group[0] or current_group[1]):
        groups.append(current_group)
        print(f"\n  Final incomplete group added as Group {len(groups)}")
        print(f"    Opposite 1: {current_group[0]}")
        print(f"    Opposite 2: {current_group[1]}")
    
    print(f"\n{'='*60}")
    print(f"GROUPING COMPLETE: {len(groups)} groups created for {shape_name}")
    print(f"{'='*60}\n")
    
    return groups

def calculate_ic_count_from_sides(sides_dict: Dict) -> int:
    """
    Calculate IC requirement by counting 'IC' occurrences in sides and dividing by 2.
    
    Args:
        sides_dict: Dictionary of sides {side1: value, side2: value, ...}
    
    Returns:
        IC count (total IC occurrences / 2)
    """
    ic_count = 0
    
    for key, value in sides_dict.items():
        side_text = str(value).upper()
        # Count occurrences of 'IC' in this side
        # Use word boundary to avoid counting IC within other words
        ic_occurrences = len(re.findall(r'\bIC\b', side_text))
        ic_count += ic_occurrences
    
    # Divide by 2 as per requirement
    return ic_count // 2

def calculate_parallel_panels_7step(opp1_parts: List[str], opp2_parts: List[str]) -> Tuple[List[List], List[List]]:
    """
    Implements the exact 7-Step Optimization Process with detailed debug logging.
    Treats every part as an individual container and calculates strictly.
    """
    print(f"\n--- START 7-STEP CALCULATION ---")
    print(f"Opposite 1 Input: {opp1_parts}")
    print(f"Opposite 2 Input: {opp2_parts}")

    class PartState:
        def __init__(self, part_str, side_id, idx, global_start):
            self.side_id = side_id 
            self.part_idx = idx
            self.id_str = f"Opp{side_id}-P{idx+1}"
            self.original_str = str(part_str).strip()
            s = self.original_str.upper()

            nums = re.findall(r'\d+', s)
            self.raw_len = int(nums[0]) if nums else 0
            
            # IC detection
            ic_matches = list(re.finditer(r'\bIC\b', s))
            self.ic_count = len(ic_matches)
            self.has_ic_start = False
            self.has_ic_end = False
            
            if self.ic_count > 0:
                num_pos = s.find(str(self.raw_len))
                for m in ic_matches:
                    if m.start() < num_pos: 
                        self.has_ic_start = True
                    if m.start() > num_pos: 
                        self.has_ic_end = True

            # Step 1 effective length
            self.effective_len = self.raw_len - (self.ic_count * 100)
            self.current_len = self.effective_len
            
            self.global_start = global_start
            self.global_end = global_start + self.raw_len
            
            self.panels_head = []
            self.panels_fill = []
            self.panels_tail = []
            
        def add_tail(self, size, p_type):
            if self.current_len >= size:
                self.current_len -= size
                self.panels_tail.append({'size': size, 'type': p_type})
            else:
                if self.current_len > 0:
                    self.panels_tail.append({'size': self.current_len, 'type': p_type})
                    self.current_len = 0

        def add_match(self, size, p_type):
            if self.current_len >= size:
                self.current_len -= size
                self.panels_head.append({'size': size, 'type': p_type})
            else:
                if self.current_len > 0:
                    self.panels_head.append({'size': self.current_len, 'type': p_type})
                    self.current_len = 0

        def __repr__(self):
            p_list = [p['size'] for p in self.panels_head + self.panels_fill + self.panels_tail]
            base = [str(self.current_len)] if self.current_len > 0 else []
            if self.has_ic_start: base.insert(0, "IC")
            if self.has_ic_end: base.append("IC")
            return f"{self.original_str} -> Rem:{self.current_len} P:{p_list}"

    # --- Step 1: Initialization ---
    s1_parts = []
    curr = 0
    for i, p in enumerate(opp1_parts):
        part = PartState(p, 1, i, curr)
        s1_parts.append(part)
        curr += part.raw_len
        
    s2_parts = []
    curr = 0
    for i, p in enumerate(opp2_parts):
        part = PartState(p, 2, i, curr)
        s2_parts.append(part)
        curr += part.raw_len

    print(f"\nStep 1 Result (Effective Lengths Calculated):")
    for p in s1_parts + s2_parts:
        print(f"  {p.id_str}: {p.original_str} -> Effective: {p.effective_len}")

    # Helper: find opposite part
    def find_target_part(parts_list, source_part, position_check='start'):
        check_pos = source_part.global_start if position_check == 'start' else source_part.global_end - 1        
        for p in parts_list:
            if p.global_start <= check_pos < p.global_end:
                return p
        idx = min(source_part.part_idx, len(parts_list)-1)
        return parts_list[idx]

    # --- Step 2: Generate IC 100s ---
    for p1 in s1_parts:
        if p1.has_ic_start:
            find_target_part(s2_parts, p1, 'start').add_tail(100, 'ic_gen')
        if p1.has_ic_end:
            find_target_part(s2_parts, p1, 'end').add_tail(100, 'ic_gen')

    for p2 in s2_parts:
        if p2.has_ic_start:
            find_target_part(s1_parts, p2, 'start').add_tail(100, 'ic_gen')
        if p2.has_ic_end:
            find_target_part(s1_parts, p2, 'end').add_tail(100, 'ic_gen')

    print(f"\nStep 2 Result (Cross-Generated 100s):")
    print(f"  Opposite 1: {[p.__repr__() for p in s1_parts]}")
    print(f"  Opposite 2: {[p.__repr__() for p in s2_parts]}")

    # --- Step 3: Mirror small segments ---
    for i, p1 in enumerate(s1_parts):
        if 0 < p1.effective_len <= 600 and p1.current_len > 0:
            mirror_size = p1.current_len
            p1.add_match(mirror_size, 'mirror_self')
            mid = (p1.global_start + p1.global_end) / 2
            target = None
            for t in s2_parts:
                if t.global_start <= mid < t.global_end:
                    target = t
                    break
            if not target:
                target = s2_parts[min(i, len(s2_parts)-1)]
            target.add_tail(mirror_size, 'mirror')

    for i, p2 in enumerate(s2_parts):
        if 0 < p2.effective_len <= 600 and p2.current_len > 0:
            mirror_size = p2.current_len
            p2.add_match(mirror_size, 'mirror_self')
            mid = (p2.global_start + p2.global_end) / 2
            target = None
            for t in s1_parts:
                if t.global_start <= mid < t.global_end:
                    target = t
                    break
            if not target:
                target = s1_parts[min(i, len(s1_parts)-1)]
            target.add_tail(mirror_size, 'mirror')

    print(f"\nStep 3 Result (Mirrors):")
    print(f"  Opposite 1: {[p.__repr__() for p in s1_parts]}")
    print(f"  Opposite 2: {[p.__repr__() for p in s2_parts]}")

    # --- Step 4 & 5: Smallest-first matching ---
    print(f"\nStep 4 & 5: Match Smallest Dimension:")
    
    while True:
        active_s1 = [p for p in s1_parts if p.current_len > 0]
        active_s2 = [p for p in s2_parts if p.current_len > 0]
        
        if not active_s1 and not active_s2:
            break
        
        if not active_s1:
            for p in active_s2:
                p.add_match(p.current_len, 'fill')
            continue

        if not active_s2:
            for p in active_s1:
                p.add_match(p.current_len, 'fill')
            continue

        min_len = float('inf')
        source_side = 0
        source_part = None
        
        for p in active_s1:
            if p.current_len < min_len:
                min_len = p.current_len
                source_side = 1
                source_part = p
        
        for p in active_s2:
            if p.current_len < min_len:
                min_len = p.current_len
                source_side = 2
                source_part = p
                
        print(f"  Smallest found: {min_len} on Opp{source_side} ({source_part.id_str})")
        source_part.add_match(min_len, 'match_self')

        targets = active_s2 if source_side == 1 else active_s1
        target = targets[0]
        target.add_match(min_len, 'match')

        print(f"  Matched: {source_part.id_str} and {target.id_str} with {min_len}")

    print(f"\nStep 4 & 5 Result:")
    print(f"  Opposite 1: {[p.__repr__() for p in s1_parts]}")
    print(f"  Opposite 2: {[p.__repr__() for p in s2_parts]}")

    # --- Step 6 & 7: Finalize with MULTIPLE MERGES ---
    print(f"\nStep 6 & 7: Standardize & Merge")

    def finalize_part(part):
        body_panels = part.panels_head + part.panels_fill
        tail_panels = part.panels_tail
        
        std_body = []
        for p in body_panels:
            size = p['size']
            if size > 600:
                while size > 600:
                    std_body.append({'size': 600, 'type': 'std'})
                    size -= 600
                if size > 0:
                    std_body.append({'size': size, 'type': 'std'})
            else:
                std_body.append(p)

        std_body.sort(key=lambda x: x['size'], reverse=True)
        work_list = std_body + tail_panels
        
        # ----- MULTI-MERGE LOGIC -----
        merged_list = []
        i = 0
        while i < len(work_list):
            curr = work_list[i]
            val = curr['size']
            curr_is_ic = (curr.get('type') == 'ic_gen')

            j = i + 1
            while j < len(work_list):
                nxt = work_list[j]
                nxt_is_ic = (nxt.get('type') == 'ic_gen')

                if (curr_is_ic or nxt_is_ic) and (val + nxt['size'] <= 600):
                    print(f"  {part.id_str}: Merging {val} + {nxt['size']} -> {val + nxt['size']}")
                    val += nxt['size']
                    j += 1
                else:
                    break

            merged_list.append(val)
            i = j
        # ----- END MULTI-MERGE LOGIC -----

        final = []
        if part.has_ic_start: final.append("IC")
        final.extend(merged_list)
        if part.has_ic_end: final.append("IC")
        return final

    s1_res = [finalize_part(p) for p in s1_parts]
    s2_res = [finalize_part(p) for p in s2_parts]

    print(f"\nFinal Result:")
    print(f"  Opposite 1: {s1_res}")
    print(f"  Opposite 2: {s2_res}")
    print(f"--- END CALCULATION ---\n")
    
    return s1_res, s2_res


def optimize_panels_and_accessories(castings: List[Casting], primary_idx: int) -> Dict:
    print("\nOptimizing panel layouts with 7-Step Parallel Logic...")
    
    ordered_castings = castings
    panel_inventory = {}
    ic_inventory = 0
    ec_inventory = 0
    new_panels_added = {}
    new_ic_added = 0
    new_ec_added = 0
    panels_used_per_casting = []
    ic_used_per_casting = []
    ec_used_per_casting = []
    total_joins = 0
    joins_per_casting = []

    for cast_idx, casting in enumerate(ordered_castings):
        casting_panels_used = {}
        casting_ic_used = casting.get_total_ic_requirement()
        casting_ec_used = casting.get_total_ec_requirement()
        casting_joins = 0
        
        print(f"\n--- Processing {casting.name} ---")
        
        for shape in casting.shapes:
            sides_filled = set()
            
            if hasattr(shape, 'grouped_sides') and shape.grouped_sides:
                for group in shape.grouped_sides:
                    opp1_parts = group[0]
                    opp2_parts = group[1]
                    
                    # Calculate 7-Step
                    p1_layout_lists, p2_layout_lists = calculate_parallel_panels_7step(opp1_parts, opp2_parts)
                    
                    # Map back to sides
                    def map_results(parts_list, result_lists):
                        res_idx = 0
                        for i, side_val in enumerate(shape.sides):
                            if i in sides_filled: continue
                            if res_idx >= len(parts_list): break
                            
                            # Loose matching by length (as strings might differ slightly)
                            s_nums = re.findall(r'\d+', str(side_val))
                            p_nums = re.findall(r'\d+', str(parts_list[res_idx]))
                            s_len = int(s_nums[0]) if s_nums else 0
                            p_len = int(p_nums[0]) if p_nums else 0
                            
                            if s_len == p_len:
                                shape.panel_layout[i] = result_lists[res_idx]
                                sides_filled.add(i)
                                res_idx += 1
                    
                    map_results(opp1_parts, p1_layout_lists)
                    map_results(opp2_parts, p2_layout_lists)

            # Fallback for non-grouped sides
            for i, side in enumerate(shape.sides):
                if i not in sides_filled:
                    shape.panel_layout[i] = get_panel_combination_exact(side)

            # Inventory Process
            for layout in shape.panel_layout:
                for item in layout:
                    if item == "IC": continue
                    panel_size = int(item)
                    
                    success, consumed = try_reuse_or_join(panel_size, panel_inventory)
                    if success:
                        if len(consumed) > 1:
                            casting_joins += 1
                            total_joins += 1
                    else:
                        new_panels_added[panel_size] = new_panels_added.get(panel_size, 0) + 1
                    
                    casting_panels_used[panel_size] = casting_panels_used.get(panel_size, 0) + 1

        print(f"  Joins performed: {casting_joins}")
        joins_per_casting.append(casting_joins)
        
        # Balance Accessories
        if ic_inventory >= casting_ic_used: ic_inventory -= casting_ic_used
        else:
            new_ic_added += (casting_ic_used - ic_inventory)
            ic_inventory = 0
            
        if ec_inventory >= casting_ec_used: ec_inventory -= casting_ec_used
        else:
            new_ec_added += (casting_ec_used - ec_inventory)
            ec_inventory = 0
            
        # Return to Inventory
        for p, c in casting_panels_used.items():
            panel_inventory[p] = panel_inventory.get(p, 0) + c
        ic_inventory += casting_ic_used
        ec_inventory += casting_ec_used
        
        panels_used_per_casting.append(casting_panels_used)
        ic_used_per_casting.append(casting_ic_used)
        ec_used_per_casting.append(casting_ec_used)

    return {
        "new_panels_added": new_panels_added,
        "new_ic_added": new_ic_added,
        "new_ec_added": new_ec_added,
        "panels_used_per_casting": panels_used_per_casting,
        "ic_used_per_casting": ic_used_per_casting,
        "ec_used_per_casting": ec_used_per_casting,
        "panel_inventory": panel_inventory,
        "total_joins": total_joins,
        "joins_per_casting": joins_per_casting
    }

def print_results_with_accessories(castings: List[Casting], primary_idx: int, optimization_results: Dict) -> None:
    print(f"\nResults (Primary Casting: {castings[primary_idx].name})\n")

    panels_used_per_casting = optimization_results["panels_used_per_casting"]
    ic_used_per_casting = optimization_results["ic_used_per_casting"]
    ec_used_per_casting = optimization_results["ec_used_per_casting"]
    new_panels_added = optimization_results["new_panels_added"]
    new_ic_added = optimization_results["new_ic_added"]
    new_ec_added = optimization_results["new_ec_added"]
    total_joins = optimization_results.get("total_joins", 0)
    joins_per_casting = optimization_results.get("joins_per_casting", [])

    for i, casting in enumerate(castings):
        print(f"{'*' * 20} {casting.name} {'*' * 20}")
        print("PRIMARY" if i == primary_idx else f"SECONDARY #{i if i > primary_idx else i+1}")
        
        for shape in casting.shapes:
            print(f"\n  Shape: {shape.name}")
            print(f"    IC Requirement: {shape.ic_requirement}, EC Requirement: {shape.ec_requirement}")
            
            if hasattr(shape, 'grouped_sides') and shape.grouped_sides:
                print(f"    Side Groups:")
                for group_idx, group in enumerate(shape.grouped_sides, 1):
                    print(f"      Group {group_idx}:")
                    print(f"        Opposite 1: {group[0]}")
                    print(f"        Opposite 2: {group[1]}")
            
            for side_idx, side_length in enumerate(shape.sides):
                panels = shape.panel_layout[side_idx]
                
                # Correct sum calc
                numeric_sum = sum(p for p in panels if isinstance(p, (int, float)))
                ic_in_panel = panels.count("IC")
                total_covered = numeric_sum + (ic_in_panel * 100)
                
                print(f"    Side {side_idx+1} (Length: {side_length}): {panels}")
                
                if total_covered != side_length:
                     print(f"      ⚠️ WARNING: Sum mismatch! Expected {side_length}, got {total_covered}")
        
        # Statistics printing (kept concise for the update)
        print(f"  Panels used: {panels_used_per_casting[i]}")
        print(f"  IC: {ic_used_per_casting[i]}, EC: {ec_used_per_casting[i]}")

    print("\n" + "=" * 50)
    print("REQUIREMENTS BY CASTING")
    print("=" * 50)
    
    # Calculate requirements per casting
    casting_requirements = []
    cumulative_panels = {}
    cumulative_ic = 0
    cumulative_ec = 0
    
    for i, casting in enumerate(castings):
        used_panels = panels_used_per_casting[i]
        used_ic = ic_used_per_casting[i]
        used_ec = ec_used_per_casting[i]
        
        if i == primary_idx:
            new_panels = sum(used_panels.values())
            new_ic = used_ic
            new_ec = used_ec
            reused_panels = 0
            reused_ic = 0
            reused_ec = 0
        else:
            new_panels = 0
            reused_panels = 0
            temp_inventory = cumulative_panels.copy()
            
            for panel, count in used_panels.items():
                reused = min(temp_inventory.get(panel, 0), count)
                reused_panels += reused
                new_panels += count - reused
                temp_inventory[panel] = temp_inventory.get(panel, 0) - reused
            
            reused_ic = min(cumulative_ic, used_ic)
            new_ic = max(0, used_ic - cumulative_ic)
            reused_ec = min(cumulative_ec, used_ec)
            new_ec = max(0, used_ec - cumulative_ec)
        
        casting_requirements.append({
            'name': casting.name,
            'is_primary': i == primary_idx,
            'total_panels': sum(used_panels.values()),
            'total_ic': used_ic,
            'total_ec': used_ec,
            'new_panels': new_panels,
            'new_ic': new_ic,
            'new_ec': new_ec,
            'reused_panels': reused_panels,
            'reused_ic': reused_ic,
            'reused_ec': reused_ec
        })
        
        casting_type = "PRIMARY" if i == primary_idx else "SECONDARY"
        print(f"{casting.name} ({casting_type}): {new_panels} panels, {new_ic} IC, {new_ec} EC needed")
        
        for panel, count in used_panels.items():
            cumulative_panels[panel] = cumulative_panels.get(panel, 0) + count
        cumulative_ic += used_ic
        cumulative_ec += used_ec
    
    total_new_panels = sum(req['new_panels'] for req in casting_requirements)
    total_new_ic = sum(req['new_ic'] for req in casting_requirements)
    total_new_ec = sum(req['new_ec'] for req in casting_requirements)
    
    print(f"\nTotal new resources needed for entire project:")
    print(f"  Panels: {total_new_panels}")
    print(f"  IC: {total_new_ic}")
    print(f"  EC: {total_new_ec}")
    print(f"\nPanel joining statistics:")
    print(f"  Total joins performed: {total_joins}")
    print(f"  Joins saved {total_joins} new panel(s) from being fabricated")

    total_panels_used = sum(req['total_panels'] for req in casting_requirements)
    total_ic_used = sum(req['total_ic'] for req in casting_requirements)
    total_ec_used = sum(req['total_ec'] for req in casting_requirements)
    
    total_panels_reused = sum(req['reused_panels'] for req in casting_requirements)
    total_ic_reused = sum(req['reused_ic'] for req in casting_requirements)
    total_ec_reused = sum(req['reused_ec'] for req in casting_requirements)
    
    total_resources_used = total_panels_used + total_ic_used + total_ec_used
    total_resources_reused = total_panels_reused + total_ic_reused + total_ec_reused
    
    if total_resources_used > 0:
        overall_efficiency = (total_resources_reused / total_resources_used) * 100
        panel_efficiency = (total_panels_reused / total_panels_used) * 100 if total_panels_used > 0 else 0
        ic_efficiency = (total_ic_reused / total_ic_used) * 100 if total_ic_used > 0 else 0
        ec_efficiency = (total_ec_reused / total_ec_used) * 100 if total_ec_used > 0 else 0
        
        print(f"\nOverall resource reuse efficiency: {overall_efficiency:.1f}%")
        print(f"Panel reuse efficiency: {panel_efficiency:.1f}%")
        print(f"IC reuse efficiency: {ic_efficiency:.1f}%")
        print(f"EC reuse efficiency: {ec_efficiency:.1f}%")
    
    secondary_castings = [req for req in casting_requirements if not req['is_primary']]
    if len(secondary_castings) > 0:
        secondary_panels_used = sum(req['total_panels'] for req in secondary_castings)
        secondary_ic_used = sum(req['total_ic'] for req in secondary_castings)
        secondary_ec_used = sum(req['total_ec'] for req in secondary_castings)
        
        secondary_panels_reused = sum(req['reused_panels'] for req in secondary_castings)
        secondary_ic_reused = sum(req['reused_ic'] for req in secondary_castings)
        secondary_ec_reused = sum(req['reused_ec'] for req in secondary_castings)
        
        secondary_resources_used = secondary_panels_used + secondary_ic_used + secondary_ec_used
        secondary_resources_reused = secondary_panels_reused + secondary_ic_reused + secondary_ec_reused
        
        if secondary_resources_used > 0:
            secondary_overall_efficiency = (secondary_resources_reused / secondary_resources_used) * 100
            secondary_panel_efficiency = (secondary_panels_reused / secondary_panels_used) * 100 if secondary_panels_used > 0 else 0
            secondary_ic_efficiency = (secondary_ic_reused / secondary_ic_used) * 100 if secondary_ic_used > 0 else 0
            secondary_ec_efficiency = (secondary_ec_reused / secondary_ec_used) * 100 if secondary_ec_used > 0 else 0
            
            print(f"\nSecondary resource reuse efficiency: {secondary_overall_efficiency:.1f}%")
            print(f"Secondary panel reuse efficiency: {secondary_panel_efficiency:.1f}%")
            print(f"Secondary IC reuse efficiency: {secondary_ic_efficiency:.1f}%")
            print(f"Secondary EC reuse efficiency: {secondary_ec_efficiency:.1f}%")

def convert_extracted_data_to_castings(castings_data: List[Dict]) -> List[Casting]:
    """
    UPDATED: Preserves raw side strings so 7-step logic can detect 'IC' text.
    """
    castings = []
    
    for casting_data in castings_data:
        casting = Casting(casting_data["name"])
        
        for shape_name, shape_info in casting_data["shapes"].items():
            raw_sides = shape_info["sides"]
            ec_count = shape_info.get("EC", 0)
            
            # Calculate IC count
            ic_count = calculate_ic_count_from_sides(raw_sides)
            
            # Group sides into opposites
            grouped_sides = group_sides_into_opposites(shape_name, raw_sides)
            
            # Extract actual numeric lengths for the Shape object (for total length calcs)
            # BUT we keep the raw strings in `grouped_sides` for the optimizer
            actual_lengths = []
            
            # Use raw_sides dict to populate the base numeric sides list
            sorted_keys = sorted(raw_sides.keys(), key=lambda x: int(re.findall(r'\d+', x)[0]))
            for key in sorted_keys:
                val = raw_sides[key]
                # Convert to int for the Shape.sides list
                if isinstance(val, (int, float)):
                    actual_lengths.append(int(val))
                else:
                    # Extract number from string
                    nums = re.findall(r'\d+', str(val))
                    actual_lengths.append(int(nums[0]) if nums else 0)

            # Create shape
            shape = Shape(shape_name, actual_lengths, ic_count=ic_count, ec_count=ec_count)
            shape.grouped_sides = grouped_sides  # IMPORTANT: Contains raw strings like "2800 IC"
            casting.add_shape(shape)
        
        castings.append(casting)
    
    return castings

def export_to_excel_with_accessories(panel_data, ic_count, ec_count, castings, primary_idx, optimization_results):
    """
    Export panels and accessories to Excel - TWO SHEETS.
    Sheet 1: Consolidated Material (Styled to match Sheet 2)
    Sheet 2: Casting Wise Material Breakdown (Existing Logic)
    """
    
    # Create workbook
    wb = Workbook()
    
    # Define Styles (Shared between sheets for consistency)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow for totals if needed
    
    # ==============================================================================
    # SHEET 1: CONSOLIDATED MATERIAL QUANTITY (STYLED)
    # ==============================================================================
    ws1 = wb.active
    ws1.title = "Consolidated Material"
    
    # --- Header Section ---
    ws1.merge_cells('A1:H1')
    ws1['A1'] = "COSMOS CONSTRUCTION MACHINERIES & EQUIPMENT PVT.LTD."
    ws1['A1'].font = Font(size=14, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:H2')
    ws1['A2'] = "Plot No.E/307/5, Gat No.307, Nanekarwadi, Next to MINDA, Chakan, Pune. Phone -+91 86008 84281"
    ws1['A2'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A3:H3')
    ws1['A3'] = "Email: aluformservice@cosmossales.com"
    ws1['A3'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A5:H5')
    ws1['A5'] = "TOTAL CONSOLIDATED MATERIAL QUANTITY"
    ws1['A5'].alignment = Alignment(horizontal='center')
    ws1['A5'].font = Font(bold=True)

    ws1.merge_cells('A6:H6')
    ws1['A6'] = "SHUTTERING AREA STATEMENT"
    ws1['A6'].alignment = Alignment(horizontal='center')
    ws1['A6'].font = Font(bold=True)

    ws1.merge_cells('A7:G7')
    ws1['A7'] = "Project Name"
    ws1['A7'].font = Font(bold=True)
    ws1['A7'].alignment = Alignment(horizontal='center')

    ws1['H7'] = datetime.now().strftime("%d.%m.%Y")
    ws1['H7'].font = Font(bold=True)
    ws1['H7'].alignment = Alignment(horizontal='right')

    # --- Table Headers ---
    sorted_panels = sorted(panel_data, key=lambda x: x["width"], reverse=True)
    table_headers = [
        "Sr. No.", "Panel Size (mm)", "Width (mm)", "Code", "Length (mm)",
        "Total No. Of Panel Quantity", "Unit Area of Panel in Sq.m", "Total Area in Sq.m of all Panels"
    ]
    ws1.append(table_headers)
    
    # Style the Headers (Row 8)
    for col in range(1, len(table_headers) + 1):
        cell = ws1.cell(row=8, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill

    # --- Add panel data ---
    total_area_sum = 0
    total_qty_sum = 0
    row_num = 9
    
    for idx, panel in enumerate(sorted_panels, start=1):
        width = panel["width"]
        length = panel["length"]
        code = panel["code"]
        qty = panel["quantity"]
        panel_size_str = f"{width}{code}{length}"
        unit_area = round((width * length) / 1_000_000, 3)
        total_area = round(unit_area * qty, 2)
        total_area_sum += total_area
        total_qty_sum += qty
        
        ws1.append([
            idx,
            panel_size_str,
            width,
            code,
            length,
            qty,
            unit_area,
            total_area
        ])
        
        # Style Data Rows
        for col in range(1, 9):
            cell = ws1.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        row_num += 1
    
    # --- Add IC accessories ---
    if ic_count > 0:
        ws1.append([
            row_num - 8,
            "IC (Internal Corner)",
            "-",
            "IC",
            "-",
            ic_count,
            "-",
            "-"
        ])
        # Style IC Row
        for col in range(1, 9):
            cell = ws1.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        row_num += 1
        total_qty_sum += ic_count
    
    # --- Add EC accessories ---
    if ec_count > 0:
        ws1.append([
            row_num - 8,
            "EC (External Corner)",
            "-",
            "EC",
            "-",
            ec_count,
            "-",
            "-"
        ])
        # Style EC Row
        for col in range(1, 9):
            cell = ws1.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        row_num += 1
        total_qty_sum += ec_count

    # --- Add Totals Row ---
    ws1.append(["", "", "", "", "", total_qty_sum])
    ws1.cell(row=ws1.max_row, column=6).font = Font(bold=True)
    
    # Style the first total row
    for col in range(1, 9):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col == 6: # Highlight the Qty
             cell.fill = total_fill
    
    row_num += 1
    
    # Add Area Total Row
    ws1.append(["", "", "", "", "", "", "Total Shuttering Area in Sq.m", round(total_area_sum, 2)])
    ws1.cell(row=ws1.max_row, column=7).font = Font(bold=True)
    ws1.cell(row=ws1.max_row, column=8).font = Font(bold=True)
    
    # Style the area total row
    for col in range(1, 9):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col in [7, 8]: # Highlight the Area labels
             cell.fill = total_fill

    # Set column widths
    for col in range(1, 9):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # ==============================================================================
    # SHEET 2: CASTING WISE MATERIAL BREAKDOWN (EXISTING LOGIC - NO CHANGES)
    # ==============================================================================
    ws2 = wb.create_sheet(title="Casting Wise Breakdown")
    
    # --- Company Header ---
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "COSMOS CONSTRUCTION MACHINERIES & EQUIPMENT PVT.LTD."
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('A2:G2')
    ws2['A2'] = "Plot No.E/307/5, Gat No.307, Nanekarwadi, Next to MINDA, Chakan, Pune. Phone -+91 86008 84281"
    ws2['A2'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('A3:G3')
    ws2['A3'] = "Email: aluformservice@cosmossales.com"
    ws2['A3'].alignment = Alignment(horizontal='center')

    # --- Sheet Title ---
    ws2.merge_cells('A5:G5')
    ws2['A5'] = "CASTING WISE MATERIAL BREAKDOWN"
    ws2['A5'].alignment = Alignment(horizontal='center')
    ws2['A5'].font = Font(bold=True, size=12)

    # --- Date ---
    ws2.merge_cells('A6:F6')
    ws2['G6'] = datetime.now().strftime("%d.%m.%Y")
    ws2['G6'].font = Font(bold=True)
    ws2['G6'].alignment = Alignment(horizontal='center')

    current_row = 7

    # --- Loop through Castings ---
    for casting_idx, casting in enumerate(castings, start=1):
        
        # 1. Casting Header (Centered)
        ws2.merge_cells(f'A{current_row}:G{current_row}')
        ws2[f'A{current_row}'] = f"Casting-{casting_idx}"
        ws2[f'A{current_row}'].font = Font(bold=True)
        ws2[f'A{current_row}'].alignment = Alignment(horizontal='center')
        ws2[f'A{current_row}'].fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for col in range(1, 8):
            ws2.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1

        # 2. Column Headers
        headers = ["Sr. no.", "Wall Name", "Panel Size (mm)", "Width", "Code", "Length", "Total No. of Panels"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws2.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            cell.fill = header_fill
        
        current_row += 1
        
        # 3. Wall Data
        sr_no = 1
        for shape in casting.shapes:
            wall_panels = {}
            
            # Count standard panels
            for side_layout in shape.panel_layout:
                for item in side_layout:
                    if isinstance(item, (int, float)):
                        key = (int(item), "WS", 2400)
                        wall_panels[key] = wall_panels.get(key, 0) + 1
            
            # Add ICs
            if shape.ic_requirement > 0:
                key = (200, "IC", 2400)
                wall_panels[key] = wall_panels.get(key, 0) + shape.ic_requirement
                
            # Add ECs
            if shape.ec_requirement > 0:
                key = (130, "EC", 2400)
                wall_panels[key] = wall_panels.get(key, 0) + shape.ec_requirement

            # Sort
            def sort_key(k):
                width, code, _ = k
                if code == "WS": return (0, -width)
                if code == "IC": return (1, 0)
                if code == "EC": return (2, 0)
                return (3, 0)

            sorted_keys = sorted(wall_panels.keys(), key=sort_key)

            # Print rows
            for (width, code, length) in sorted_keys:
                count = wall_panels[(width, code, length)]
                if code == "IC":
                    panel_size_str = "(100+100) IC 2400"
                elif code == "EC":
                    panel_size_str = "(65+65) EC 2400"
                else:
                    panel_size_str = f"{width} WS {length}"

                row_data = [sr_no, shape.name, panel_size_str, width, code, length, count]
                
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws2.cell(row=current_row, column=col_idx)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                current_row += 1
                sr_no += 1
        
        current_row += 1

    # Adjust Column Widths for Sheet 2
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 20

    # Save workbook
    wb.save("consolidated_list.xlsx")
    print("✅ Excel file 'consolidated_list.xlsx' generated successfully.")
    
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/api/extract-castings', methods=['POST'])
def extract_castings_route():
    if 'pdf' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['pdf']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        try:
            file.save(file_path)
            
            # Extract castings using text-based extraction with IC/EC detection
            castings_data = extract_castings_from_pdf(file_path)
            
            # Display extracted data for verification
            print("\n" + "=" * 80)
            print("EXTRACTED CASTING DATA (Text-Based Extraction with IC/EC)")
            print("=" * 80)
            print(json.dumps(castings_data, indent=2))
            print("=" * 80 + "\n")
            
            os.remove(file_path)
            return jsonify(castings_data), 200
        except Exception as e:
            logging.exception("Failed to process PDF")
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({'error': f"Failed to process PDF: {str(e)}"}), 500
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/api/optimize', methods=['POST']) 
def optimize():
    try:
        data = request.get_json()
        if not data or 'castings' not in data or 'primaryIdx' not in data:
            return jsonify({'error': 'Invalid request data'}), 400

        # Display received casting data for verification
        print("\n" + "=" * 80)
        print("RECEIVED CASTING DATA FOR OPTIMIZATION")
        print("=" * 80)
        print(json.dumps(data['castings'], indent=2))
        print("=" * 80 + "\n")

        castings = convert_extracted_data_to_castings(data['castings'])
        primary_idx = int(data['primaryIdx'])

        if primary_idx < 0 or primary_idx >= len(castings):
            return jsonify({'error': f"Invalid primaryIdx: {primary_idx}"}), 400

        # Analyze castings
        analysis_results = analyze_castings(castings)

        # Perform optimization with accessories
        optimization_results = optimize_panels_and_accessories(castings, primary_idx)

        # Capture print_results output
        original_stdout = sys.stdout
        sys.stdout = buffer = StringIO()
        print_results_with_accessories(castings, primary_idx, optimization_results)
        sys.stdout = original_stdout
        text_summary = buffer.getvalue()

        # Prepare panel_data for Excel export
        new_panels_added = optimization_results["new_panels_added"]
        new_ic_added = optimization_results["new_ic_added"]
        new_ec_added = optimization_results["new_ec_added"]
        
        panel_data = [
            {
                "width": panel_size,
                "length": 2400,
                "code": "WS",
                "quantity": quantity
            }
            for panel_size, quantity in new_panels_added.items()
        ]

        # Export to Excel with accessories
        export_to_excel_with_accessories(panel_data, new_ic_added, new_ec_added, castings, primary_idx, optimization_results)

        # Calculate efficiency metrics
        panels_used_per_casting = optimization_results["panels_used_per_casting"]
        ic_used_per_casting = optimization_results["ic_used_per_casting"]
        ec_used_per_casting = optimization_results["ec_used_per_casting"]
        
        # Calculate requirements per casting
        casting_requirements = []
        cumulative_panels = {}
        cumulative_ic = 0
        cumulative_ec = 0
        
        for i, casting in enumerate(castings):
            used_panels = panels_used_per_casting[i]
            used_ic = ic_used_per_casting[i]
            used_ec = ec_used_per_casting[i]
            
            if i == primary_idx:
                new_panels = sum(used_panels.values())
                new_ic = used_ic
                new_ec = used_ec
                reused_panels = 0
                reused_ic = 0
                reused_ec = 0
            else:
                new_panels = 0
                reused_panels = 0
                temp_inventory = cumulative_panels.copy()
                
                for panel, count in used_panels.items():
                    reused = min(temp_inventory.get(panel, 0), count)
                    reused_panels += reused
                    new_panels += count - reused
                    temp_inventory[panel] = temp_inventory.get(panel, 0) - reused
                
                reused_ic = min(cumulative_ic, used_ic)
                new_ic = max(0, used_ic - cumulative_ic)
                reused_ec = min(cumulative_ec, used_ec)
                new_ec = max(0, used_ec - cumulative_ec)
            
            casting_requirements.append({
                'name': casting.name,
                'is_primary': i == primary_idx,
                'total_panels': sum(used_panels.values()),
                'total_ic': used_ic,
                'total_ec': used_ec,
                'new_panels': new_panels,
                'new_ic': new_ic,
                'new_ec': new_ec,
                'reused_panels': reused_panels,
                'reused_ic': reused_ic,
                'reused_ec': reused_ec
            })
            
            # Update cumulative inventory
            for panel, count in used_panels.items():
                cumulative_panels[panel] = cumulative_panels.get(panel, 0) + count
            cumulative_ic += used_ic
            cumulative_ec += used_ec
        
        # Calculate overall efficiency (all castings)
        total_panels_used = sum(req['total_panels'] for req in casting_requirements)
        total_ic_used = sum(req['total_ic'] for req in casting_requirements)
        total_ec_used = sum(req['total_ec'] for req in casting_requirements)
        
        total_panels_reused = sum(req['reused_panels'] for req in casting_requirements)
        total_ic_reused = sum(req['reused_ic'] for req in casting_requirements)
        total_ec_reused = sum(req['reused_ec'] for req in casting_requirements)
        
        total_resources_used = total_panels_used + total_ic_used + total_ec_used
        total_resources_reused = total_panels_reused + total_ic_reused + total_ec_reused
        
        overall_efficiency = (total_resources_reused / total_resources_used * 100) if total_resources_used > 0 else 0
        panel_efficiency = (total_panels_reused / total_panels_used * 100) if total_panels_used > 0 else 0
        ic_efficiency = (total_ic_reused / total_ic_used * 100) if total_ic_used > 0 else 0
        ec_efficiency = (total_ec_reused / total_ec_used * 100) if total_ec_used > 0 else 0
        
        # Calculate secondary efficiency (castings 2..N)
        secondary_castings = [req for req in casting_requirements if not req['is_primary']]
        
        secondary_overall_efficiency = 0
        secondary_panel_efficiency = 0
        secondary_ic_efficiency = 0
        secondary_ec_efficiency = 0
        
        if len(secondary_castings) > 0:
            secondary_panels_used = sum(req['total_panels'] for req in secondary_castings)
            secondary_ic_used = sum(req['total_ic'] for req in secondary_castings)
            secondary_ec_used = sum(req['total_ec'] for req in secondary_castings)
            
            secondary_panels_reused = sum(req['reused_panels'] for req in secondary_castings)
            secondary_ic_reused = sum(req['reused_ic'] for req in secondary_castings)
            secondary_ec_reused = sum(req['reused_ec'] for req in secondary_castings)
            
            secondary_resources_used = secondary_panels_used + secondary_ic_used + secondary_ec_used
            secondary_resources_reused = secondary_panels_reused + secondary_ic_reused + secondary_ec_reused
            
            if secondary_resources_used > 0:
                secondary_overall_efficiency = (secondary_resources_reused / secondary_resources_used) * 100
                secondary_panel_efficiency = (secondary_panels_reused / secondary_panels_used) * 100 if secondary_panels_used > 0 else 0
                secondary_ic_efficiency = (secondary_ic_reused / secondary_ic_used) * 100 if secondary_ic_used > 0 else 0
                secondary_ec_efficiency = (secondary_ec_reused / secondary_ec_used) * 100 if secondary_ec_used > 0 else 0

        # Prepare results for JSON
        results = {
            'castings': [],
            'optimization_summary': {
                'total_castings': len(castings),
                'primary_casting': castings[primary_idx].name,
                'panel_sizes_used': list(set(
                    panel for casting in castings
                    for shape in casting.shapes
                    for panels in shape.panel_layout
                    for panel in panels
                )),
                'total_new_panels_needed': sum(new_panels_added.values()),
                'total_new_ic_needed': new_ic_added,
                'total_new_ec_needed': new_ec_added,
                'total_joins_performed': optimization_results.get('total_joins', 0),
                'joins_per_casting': optimization_results.get('joins_per_casting', []),
                'overall_efficiency': {
                    'total_efficiency': round(overall_efficiency, 1),
                    'panel_efficiency': round(panel_efficiency, 1),
                    'ic_efficiency': round(ic_efficiency, 1),
                    'ec_efficiency': round(ec_efficiency, 1),
                    'total_resources_used': total_resources_used,
                    'total_resources_reused': total_resources_reused
                },
                'secondary_efficiency': {
                    'total_efficiency': round(secondary_overall_efficiency, 1),
                    'panel_efficiency': round(secondary_panel_efficiency, 1),
                    'ic_efficiency': round(secondary_ic_efficiency, 1),
                    'ec_efficiency': round(secondary_ec_efficiency, 1)
                },
                'panel_reuse_analysis': {
                    'primary_casting': {
                        'name': castings[primary_idx].name,
                        'panels_used': panels_used_per_casting[primary_idx],
                        'ic_used': ic_used_per_casting[primary_idx],
                        'ec_used': ec_used_per_casting[primary_idx],
                        'new_panels_needed': panels_used_per_casting[primary_idx],
                        'new_ic_needed': ic_used_per_casting[primary_idx],
                        'new_ec_needed': ec_used_per_casting[primary_idx]
                    },
                    'secondary_castings': []
                },
                'panel_analysis': {
                    'length_counts': analysis_results['length_counts'],
                    'preferred_sizes': analysis_results['preferred_sizes'],
                    'panel_efficiency': analysis_results['panel_efficiency']
                },
                'text_summary': text_summary
            }
        }

        # Add secondary casting analysis
        cumulative_panels = panels_used_per_casting[primary_idx].copy()
        cumulative_ic = ic_used_per_casting[primary_idx]
        cumulative_ec = ec_used_per_casting[primary_idx]
        
        for i in range(len(castings)):
            if i == primary_idx:
                continue
                
            used_panels = panels_used_per_casting[i]
            used_ic = ic_used_per_casting[i]
            used_ec = ec_used_per_casting[i]
            
            # Calculate panel reuse
            reused_panels = {}
            new_panels_needed = {}
            temp_inventory = cumulative_panels.copy()
            
            for panel, count in used_panels.items():
                reused_count = min(temp_inventory.get(panel, 0), count)
                reused_panels[panel] = reused_count
                new_panels_needed[panel] = count - reused_count
                temp_inventory[panel] = temp_inventory.get(panel, 0) - reused_count
            
            # Calculate IC/EC reuse
            reused_ic = min(cumulative_ic, used_ic)
            new_ic_needed = max(0, used_ic - cumulative_ic)
            reused_ec = min(cumulative_ec, used_ec)
            new_ec_needed = max(0, used_ec - cumulative_ec)
            
            results["optimization_summary"]["panel_reuse_analysis"]["secondary_castings"].append({
                "name": castings[i].name,
                "panels_used": used_panels,
                "ic_used": used_ic,
                "ec_used": used_ec,
                "reused_panels_from_previous": reused_panels,
                "reused_ic_from_previous": reused_ic,
                "reused_ec_from_previous": reused_ec,
                "new_panels_needed": new_panels_needed,
                "new_ic_needed": new_ic_needed,
                "new_ec_needed": new_ec_needed
            })
            
            # Update cumulative inventory
            for panel, count in used_panels.items():
                cumulative_panels[panel] = cumulative_panels.get(panel, 0) + count
            cumulative_ic += used_ic
            cumulative_ec += used_ec

        # Add casting details with IC/EC information
        for casting in castings:
            casting_data = {
                'name': casting.name,
                'total_ic_requirement': casting.get_total_ic_requirement(),
                'total_ec_requirement': casting.get_total_ec_requirement(),
                'shapes': []
            }
            for shape in casting.shapes:
                shape_data = {
                    'name': shape.name,
                    'sides': shape.sides,
                    'panel_layouts': shape.panel_layout,
                    'ic_requirement': shape.ic_requirement,
                    'ec_requirement': shape.ec_requirement
                }
                casting_data['shapes'].append(shape_data)
            results['castings'].append(casting_data)

        # Save results to file
        with open(RESULTS_FILE, 'w') as f:
            json.dump(results, f, indent=2)

        return jsonify(results), 200

    except ValueError as ve:
        logging.error(f"ValueError: {str(ve)}")
        return jsonify({'error': f"Invalid data: {str(ve)}"}), 400
    except Exception as e:
        logging.exception("Optimization failed")
        return jsonify({'error': f"Optimization failed: {str(e)}"}), 500

@app.route('/api/download-excel', methods=['GET'])
def download_excel():
    excel_file = "consolidated_list.xlsx"
    try:
        if os.path.exists(excel_file):
            return send_file(excel_file, as_attachment=True, download_name='consolidated_list.xlsx')
        else:
            return jsonify({'error': 'Excel file not found. Please run optimization first.'}), 404
    except Exception as e:
        logging.exception("Failed to download Excel file")
        return jsonify({'error': f"Failed to download Excel file: {str(e)}"}), 500

if __name__ == '__main__':
    logging.info("Starting Flask server on http://localhost:5000")
    app.run(debug=True, port=5000)